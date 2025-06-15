import numpy as np
import pandas as pd
import openpyxl
import toml
from pathlib import Path
from utilities import create_event_slices, normalize, smooth, baseline_threshold, previous_threshold, derivate_threshold
from matplotlib.figure import Figure
from typing import Optional

class SubDir:
    def __init__(self, path: Path, report_config: dict[str, str]) -> None:
        self.path = path
        self.report_path = path / f"{report_config["name"]}{path.name}{report_config["extension"]}"
        self.treatment_col_names: list[str] = []
        self.treatment_windows: dict[str, slice[int]] = {}
        self.report: Optional[pd.DataFrame] = None
        self.has_report: bool = False
        self.conditions: dict[str, str]

    def preprocessing(self, repeat: bool):
        self.parse_metadata()

        if self.report_path.exists() and not repeat:
            self.has_report = True

    def parse_events(self) -> None:
        """Reads in the event file that describes what happened during the measurement in question.

        Args:
            file (Path): The pathlib.Path object representing the event file. Must be csv.

        Returns:
            dict[str, slice[int]: This dictionary maps the names of agonists to the time windows where they were applied.
            list[str]: The list of agonist related column names for the report DataFrame.
        """
        file = self.path / "events.csv"
        events_df = pd.read_csv(file, header=0)

        treatments, start_times, stop_times = events_df["treatment"], events_df["start"], events_df["stop"]
        event_slices = create_event_slices(start_times, stop_times)
        self.treatment_windows: dict[str, slice[int]] = {k: v for k, v in zip(treatments, event_slices)}
        
        for agonist in treatments:
            if agonist == "baseline":
                continue
            self.treatment_col_names.append(agonist + "_reaction")
            self.treatment_col_names.append(agonist + "_amp")

    def parse_metadata(self) -> None:
        file = self.path / "metadata.toml"
        with open(file, "r") as f:
            metadata = toml.load(f)

        self.conditions = metadata["conditions"]

        for agonist_name, agonist_dict in metadata["treatments"].items():
            self.treatment_windows[agonist_name] = slice(agonist_dict["start"], agonist_dict["stop"])
            if agonist_name == "baseline":
                continue
            self.treatment_col_names.append(agonist_name + "_reaction")
            self.treatment_col_names.append(agonist_name + "_amp")
    
    def make_report(self, method: str, sd_multiplier: int) -> None:
        """This meant to encapsulate everything currently under the if process: block in process_subdir().
        """
        if self.has_report:
            return
                
        measurement_files = [f for f in self.path.glob("*.xlsx") if f != self.report_path]
        output_columns = ["cell_ID", "condition", "cell_type"] + self.treatment_col_names

        self.report = pd.DataFrame(columns=output_columns)
        cell_ID: int = 0
        for file in measurement_files:
            file_result = pd.DataFrame(columns=output_columns)
            if self.conditions["ratiometric_dye"]:
                cell_cols, data = self.prepare_ratiometric_data(file)
            else:
                cell_cols, data = self.prepare_non_ratiometric_data(file)
            # measurements with neurons only will be called "neuron only {number}.xlsx" whereas neuron + DPC is going to
            # be "neuron + DPC {number}.xlsx"
            condition = "N" if "only" in file.name else "N+D"
            
            # determine if cells react to each of the agonists, and how big the response amplitudes are
            # no default case because we already have a guard clause to make sure these 3 are the only options, which we
            # do in main before reading any measurement data from disk, so if the program's gonna crash it does so quickly
            match method:
                case "baseline":
                    baseline_threshold(data, self.treatment_windows, file_result, sd_multiplier)
                case "previous":
                    previous_threshold(data, self.treatment_windows, file_result, sd_multiplier)
                case "derivative":
                    derivate_threshold(data, self.treatment_windows, file_result, sd_multiplier)
            
            number_of_cells = data.shape[0]
            file_result["cell_ID"] = [x for x in range(cell_ID, cell_ID + number_of_cells)]
            cell_ID += number_of_cells
            file_result["condition"] = [condition for _ in range(number_of_cells)]
            # in the Excel files, columns will be called N1, N2, N3... for neurons and DPC1, DPC2, DPC3... for DPCs
            cell_cols = [c.strip("1234567890") for c in cell_cols]
            file_result["cell_type"] = cell_cols

            self.report = pd.concat([self.report, file_result])

        self.save_report()

    def make_graphs(self):
        measurement_files = [f for f in self.path.glob("*.xlsx") if f != self.report_path]
        if self.report is None:
            self.report = pd.read_excel(self.report_path, sheet_name="Cells")
        
        for file in measurement_files:
            graphing_path: Path = self.path / Path(file.stem)
            if not graphing_path.exists():
                Path.mkdir(graphing_path)

            ratios = pd.read_excel(file, sheet_name="py_ratios")
            cell_cols = [c for c in ratios.columns if c != "Time"]
            ratios = np.transpose(ratios.to_numpy())
            x_data, ratios = ratios[0], ratios[1:]
            reaction_cols = [col for col in self.report.columns if "_reaction" in col]

            self.graph_data(x_data.flatten(), ratios, cell_cols, self.report[reaction_cols], graphing_path)
    
    def graph_data(self, x_data: np.ndarray, traces: np.ndarray, col_names: list[str], reactions: pd.DataFrame, save_dir: Path) -> None:
        """Creates line graphs for each cell in this particular measurement file. Is called from within make_report()
        because it needs the cell trace data and that funtion only returns the report DataFrame.

        Args:
            x_data (np.ndarray): The time values as a 1d array.
            traces (np.ndarray): The ratio data to be plotted.
            treatments (dict[str, slice[int]]): The dictionary describing what agonist was used between what time points.
            Called agonist_slices elsewhere.
            reactions (pd.DataFrame): Those columns of the file result df that contain the reaction True/False values for
            each agonist used.
            save_dir (Path): The newly created directory where the graphs are supposed to be saved.
        """
        majors = [x for x in range(0, len(x_data) + 1, 60)]
        major_labels = [str(x//60) for x in majors]
        for i, (cell_name, y_data) in enumerate(zip(col_names, traces), start=0):
            fig = Figure(figsize=(10, 5))
            ax = fig.subplots(1, 1)

            ax.plot(x_data, y_data)
            ax.set_xticks(majors, labels=major_labels, minor=False)
            ax.set_xlabel("Time (min)")
            ax.set_ylabel("Ratio")

            ymin, ymax = ax.get_ylim()
            agonist_label_y = ymin - (ymax - ymin) * 0.2
            reaction_label_y = ymin - (ymax - ymin) * 0.25
            for name, time_slice in self.treatment_windows.items():
                if name == "baseline" or name == "END":
                    continue
                ax.axvline(x=time_slice.start, c="black")
                ax.axvline(x=time_slice.stop, c="black")
                ax.text(x=time_slice.start, y=agonist_label_y, s=name)
                # this next line is supposed to print TRUE under a given agonist name if the program thinks that cell reacts
                # to that agonist and FALSE otherwise
                # if name != "baseline":
                ax.text(x=time_slice.start, y=reaction_label_y, s=str(reactions.at[i, f"{name}_reaction"]).upper())

            # Cell numbering is 0 indexed on purpose!
            fig.suptitle(f"{cell_name}")
            fig.tight_layout()
            fig.savefig(save_dir / f"Cell no. {i}.png", dpi=300)
            fig.clf()
            print(f"Done with {cell_name}")

    def load_summary(self) -> None:
            self.report = pd.read_excel(self.report_path, sheet_name="Summary")

    def prepare_ratiometric_data(self, file: Path) -> tuple[list[str], np.ndarray]:
        # read in 340 and 380 data separately
        F340_data = pd.read_excel(file, sheet_name="F340")
        F380_data = pd.read_excel(file, sheet_name="F380")
        cell_cols = [c for c in F380_data.columns if c not in {"Time", "Background"}]
        
        # split the data
        x_data, bgr_380, cells_380 = F380_data["Time"].to_numpy(), F380_data["Background"].to_numpy(), F380_data[cell_cols].to_numpy()
        bgr_340, cells_340 = F340_data["Background"].to_numpy(), F340_data[cell_cols].to_numpy()
        
        # turn time and background into 2d arrays with one column each because this shape is needed for linalg.lstsq
        x_data, bgr_340, bgr_380 = x_data[:, np.newaxis], bgr_340[:, np.newaxis], bgr_380[:, np.newaxis]
        
        # substract backgrounds
        cells_340 = cells_340 - bgr_340
        cells_380 = cells_380 - bgr_380
        
        # smoothing should probably go here
        cells_340 = np.apply_along_axis(smooth, 0, cells_340)
        cells_380 = np.apply_along_axis(smooth, 0, cells_380)

        # photobleaching correction
        matrix = np.hstack((np.ones_like(x_data), x_data))
        coeffs_340, _, _, _ = np.linalg.lstsq(matrix, cells_340, rcond=None)
        coeffs_380, _, _, _ = np.linalg.lstsq(matrix, cells_380, rcond=None)
        cells_340 = cells_340 - (x_data * coeffs_340[1])
        cells_380 = cells_380 - (x_data * coeffs_380[1])
        
        # I'm working with Fura2 so the actual data of interest is the ratios between emissions at 340 and 380 nm.
        ratios = np.transpose(cells_340 / cells_380)
        self.save_processed_data(file, x_data, ratios, cell_cols)

        return cell_cols, ratios
    
    def prepare_non_ratiometric_data(self, file:Path) -> tuple[list[str], np.ndarray]:
        data = pd.read_excel(file, sheet_name="Raw")
        cell_cols = [c for c in data.columns if c not in {"Time", "Background"}]
        x_data, bgr, cells = data["Time"].to_numpy(), data["Background"].to_numpy(), data[cell_cols].to_numpy()
        x_data, bgr = x_data[:, np.newaxis], bgr[:, np.newaxis]
        
        # normalization and smoothing
        cells = np.apply_along_axis(normalize, 0, cells, baseline=self.treatment_windows["baseline"].stop)
        cells = np.apply_along_axis(smooth, 0, cells)
        
        # photobleaching correction
        matrix = np.hstack((np.ones_like(x_data), x_data))
        coeffs, _, _, _ = np.linalg.lstsq(matrix, cells, rcond=None)
        cells = cells - (x_data * coeffs[1])

        self.save_processed_data(file, x_data, cells, cell_cols)
        return cell_cols, cells.transpose()


    def save_processed_data(self, file: Path, x_data: np.ndarray, cell_data: np.ndarray, col_names: list[str]) -> None:
        col_names = ["Time"] + col_names
        data = np.vstack((x_data.flatten(), cell_data))
        data = np.transpose(data)
        if self.conditions["ratiometric_dye"]:
            sheet_name: str = "Py_ratios"
        else:
            sheet_name: str = "Processed"
        
        wb = openpyxl.load_workbook(file)
        if sheet_name in wb.sheetnames: # this is only going to be true when --repeat is used
            wb.remove(wb[sheet_name])
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        
        ws.append(col_names)
        for row in data:
            ws.append(list(row))

        wb.save(file)

    def save_report(self) -> None:
        assert self.report is not None # report is guaranteed not to be None by the time this method is called
        with pd.ExcelWriter(self.report_path) as writer:
            self.report.to_excel(writer, sheet_name="Cells", index=False)
            stats = self.report[["cell_type"] + self.treatment_col_names].value_counts()
            stats.to_excel(writer, sheet_name="Summary")